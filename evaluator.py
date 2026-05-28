import pandas as pd
import json
import re
import os
from typing import List, Dict, Set, Tuple
from sentence_transformers import SentenceTransformer, util
import numpy as np
from tqdm import tqdm

class ContentBasedRAGEvaluator:
    """
    基于内容匹配的RAG评测器
    不依赖chunk_id，通过文本相似度判断检索结果是否命中Gold Evidence
    """
    
    def __init__(self, test_excel: str, similarity_threshold: float = 0.65,
                 citation_similarity_threshold: float = 0.65,
                 auto_scoring: bool = True, scoring_model=None,
                 chroma_collection=None):
        """
        :param test_excel: 测试集Excel路径
        :param similarity_threshold: 检索评估的相似度阈值
        :param citation_similarity_threshold: 引用准确率的相似度阈值
        :param auto_scoring: 是否启用 LLM 自动评分
        :param scoring_model: 评分模型（可选，如果为 None 且 auto_scoring=True 则自动加载）
        :param chroma_collection: ChromaDB集合对象，用于根据chunk_id查询原文
        """
        self.sheets = pd.read_excel(test_excel, sheet_name=None)
        self.question_types = list(self.sheets.keys())
        self.similarity_threshold = similarity_threshold
        self.citation_similarity_threshold = citation_similarity_threshold
        self.auto_scoring = auto_scoring
        self.scoring_model = scoring_model
        self.chroma_collection = chroma_collection  # 保存向量数据库引用
        
        # 加载语义相似度模型（用于判断内容是否匹配）
        print("🔄 加载语义相似度模型...")
        
        # 尝试从 ModelScope 下载
        sim_model_name = 'BAAI/bge-m3'
        try:
            from modelscope import snapshot_download
            print(f"📥 正在通过 ModelScope 下载模型: {sim_model_name}...")
            model_path = snapshot_download(sim_model_name, cache_dir="./model_cache")
            print(f"✅ ModelScope 下载完成，缓存至: {model_path}")
            sim_model_name = model_path
        except ImportError:
            print("⚠️ 未安装 modelscope，将使用 HuggingFace 源")
        except Exception as e:
            print(f"⚠️ ModelScope 下载失败: {e}")
            print(f"   回退到 HuggingFace 源")
        
        self.sim_model = SentenceTransformer(sim_model_name)  # 中文推荐
        # 如果是英文论文，可以用: 'sentence-transformers/all-MiniLM-L6-v2'
        
        # 合并所有问题
        self.all_questions = []
        for q_type, df in self.sheets.items():
            df['question_type'] = q_type
            self.all_questions.append(df)
        self.test_df = pd.concat(self.all_questions, ignore_index=True)
        
        # 去重问题类型（防止Excel中有重复的sheet name）
        self.question_types = list(dict.fromkeys(self.sheets.keys()))
        
        print(f"✅ 初始化完成，共 {len(self.test_df)} 个问题")
        print(f"   检索评估阈值: {similarity_threshold}")
        print(f"   引用评估阈值: {citation_similarity_threshold}")
        print(f"   自动评分: {'启用' if auto_scoring else '禁用'}")


    def calculate_ragas_style_correctness(self, question: str, gold_answer: str, model_answer: str) -> Dict:
        """
        使用 LLM（Qwen2.5-14B）进行知识点拆解评分
        输出 TP/FP/FN 三个数字
        """
        if pd.isna(gold_answer) or str(gold_answer).strip() == "" or not model_answer:
            return {"f1_score": 0.0, "tp": 0, "fp": 0, "fn": 0}

        extraction_prompt = f"""你是一个公正的学术评测员。请对比【标准答案】和【模型回答】，输出三个数字。

【知识点的定义 - 重要】
知识点 = 标准答案中的【核心观点、关键概念、重要结论、主要方法/步骤】
- 一个完整的观点/定义/方法 = 1个知识点
- 同一观点下的并列概念/分类/举例 ≠ 独立知识点（算作同一个知识点的细节）
- 修饰性形容词/副词 ≠ 独立知识点（如"重要的"等）
- 举例、解释、背景说明 ≠ 独立知识点
- 论文名、作者、年份 ≠ 知识点
- 细节数据、具体数值 ≠ 独立知识点（除非是核心结论）

【评判标准】
TP（命中）：模型回答中正确提到的【标准答案中的】知识点数量
FP（幻觉）：模型回答中【捏造的、与标准答案明确矛盾的】错误信息数量
FN（遗漏）：【标准答案中有】但模型回答完全没提到的知识点数量

【FP判定规则 - 重要】
- 只有【完全捏造】或【与标准答案直接矛盾】才算FP
- 模型补充的合理细节、解释、举例 ≠ FP
- 论文名、引用来源、页码 ≠ FP
- 表述不同但意思相同 ≠ FP
- 同一概念的不同缩写/简称 ≠ FP
- 中英文表述同一概念 ≠ FP

【输出格式 - 重要】
按顺序输出三个整数：TP, FP, FN
用英文逗号分隔，例如：3,0,0
不要输出任何其他文字、标点符号或解释

【标准答案】
{gold_answer}

【模型回答】
{model_answer}

输出（三个数字，逗号分隔）："""

        try:
            messages = [
                {"role": "system", "content": "你是评测助手，只输出三个整数，用英文逗号分隔。例如：3,0,0"},
                {"role": "user", "content": extraction_prompt}
            ]
            
            # 调用 14B 模型评分（带重试）
            max_retries = 2
            raw_output = None
            
            for attempt in range(max_retries + 1):
                raw_output = self.scoring_model.generate_score(messages)
                
                # 解析：提取所有数字
                numbers = re.findall(r'\d+', raw_output)
                
                if len(numbers) >= 3:
                    tp, fp, fn = int(numbers[0]), int(numbers[1]), int(numbers[2])
                    break
                elif attempt < max_retries:
                    print(f"⚠️ 第{attempt+1}次解析失败: '{raw_output}'，重试中...")
                    messages[-1]["content"] += "\n\n【重要】只输出三个整数，用英文逗号分隔！例如：3,0,0"
                else:
                    raise ValueError(f"输出格式不正确: {raw_output}")

        except Exception as e:
            print(f"⚠️ LLM 评分失败: {e}。回退到语义相似度...")
            sim = self.compute_similarity(gold_answer, model_answer)
            return {"f1_score": sim, "tp": 0, "fp": 0, "fn": 0, "fallback": True}

        # 计算 F1 分数
        if tp == 0 and fp == 0 and fn == 0:
            f1 = 0.0
        else:
            f1 = tp / (tp + 0.5 * fp + 0.5 * fn)
            
        f1 = max(0.0, min(1.0, f1))

        return {
            "f1_score": round(f1, 3),
            "tp": tp,
            "fp": fp,
            "fn": fn
        }
    
        
    def compute_similarity(self, text1: str, text2: str) -> float:
        """
        计算两段文本的语义相似度
        """
        emb1 = self.sim_model.encode(text1, convert_to_tensor=True)
        emb2 = self.sim_model.encode(text2, convert_to_tensor=True)
        score = util.cos_sim(emb1, emb2).item()
        return score
    
    def check_retrieved_hit_gold(self, retrieved_chunk_text: str, 
                                  gold_evidence_texts: List[str]) -> Tuple[bool, float]:
        """
        判断一个检索到的chunk是否命中了任意一条Gold Evidence（使用检索阈值）
        
        :param retrieved_chunk_text: 检索到的chunk文本
        :param gold_evidence_texts: Gold Evidence文本列表（从Excel的Gold Evidence列解析）
        :return: (是否命中, 最高相似度)
        """
        if not gold_evidence_texts:
            return False, 0.0
        
        max_sim = 0.0
        for gold_text in gold_evidence_texts:
            sim = self.compute_similarity(retrieved_chunk_text, gold_text)
            max_sim = max(max_sim, sim)
        
        is_hit = max_sim >= self.similarity_threshold
        return is_hit, max_sim
    def check_citation_hit_gold(self, citation_text: str, 
                                  gold_evidence_texts: List[str]) -> Tuple[int, float]:
        if not gold_evidence_texts:
            return -1, 0.0
        
        max_sim = 0.0
        best_match_idx = -1
        
        for idx, gold_text in enumerate(gold_evidence_texts):
            sim = self.compute_similarity(citation_text, gold_text)
            if sim > max_sim:
                max_sim = sim
                best_match_idx = idx 
        
        if max_sim >= self.citation_similarity_threshold:
            return best_match_idx, max_sim
        else:
            return -1, max_sim

    def parse_gold_evidence(self, gold_evidence_str: str) -> List[str]:
        """
        解析Gold Evidence列
        
        您的测试集中Gold Evidence是文本描述，可能包含多条证据
        支持格式：
        - 单条文本: "Zhang et al. (2024) proposed..."
        - 多条文本（用换行或分号分隔）: "证据1\n证据2" 或 "证据1; 证据2"
        - JSON数组: '["证据1", "证据2"]'
        """
        if pd.isna(gold_evidence_str) or str(gold_evidence_str).strip() == "":
            return []
        
        gold_str = str(gold_evidence_str).strip()
        
        # 尝试JSON解析
        try:
            parsed = json.loads(gold_str)
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        except:
            pass
        
        # 提取双引号包裹的内容
        import re
        quoted_parts = re.findall(r'"([^"]+)"', gold_str)
        
        if quoted_parts:
            # 有双引号，提取引号内的内容，合并内部的换行
            parts = []
            for part in quoted_parts:
                # 将引号内的换行符替换为空格，合并成一条
                cleaned = ' '.join(line.strip() for line in part.split('\n') if line.strip())
                if cleaned:
                    parts.append(cleaned)
            return parts
        
        # 没有双引号，按换行或分号分隔
        if '\n' in gold_str:
            lines = gold_str.split('\n')
            # 如果每行都很短(<50字符)且行首是小写字母,说明是PDF换行,应该合并
            is_pdf_wrapping = all(
                len(line.strip()) < 50 and 
                (line.strip() == '' or line.strip()[0].islower())
                for line in lines
                if line.strip()
            )
            
            if is_pdf_wrapping:
                # PDF换行,合并成一条
                parts = [' '.join(line.strip() for line in lines)]
            else:
                # 真正的多行证据
                parts = [line.strip() for line in lines if line.strip()]
        else:
            parts = [gold_str]
        
        return [p.strip() for p in parts if p.strip()]
    
    def calculate_recall_at_k(self, reranked_chunks: List[Dict], 
                              gold_evidence_texts: List[str],
                              k_values: List[int] = [1, 3, 5, 10]) -> Dict[int, float]:
        """
        基于内容匹配计算Recall@k
        
        :param reranked_chunks: 重排后的chunk列表（按排名排序），每个包含'text'字段
        :param gold_evidence_texts: Gold Evidence文本列表
        :param k_values: 要计算的k值列表
        
        计算策略：
        - Recall@1: Binary Recall（0或1），表示第一个结果是否命中任意Gold Evidence
        - Recall@k (k>1): Coverage Recall（0到1），表示Top-K覆盖了多少比例的Gold Evidence
        """
        if not gold_evidence_texts:
            return {k: None for k in k_values}
        
        recall_results = {}
        
        for k in k_values:
            top_k_chunks = reranked_chunks[:k]
            
            if k == 1:
                # Binary Recall: 第一个结果是否命中任意一条Gold Evidence
                chunk_text = top_k_chunks[0].get('text', '') or top_k_chunks[0].get('content', '') if top_k_chunks else ''
                chunk_text = self.clean_markdown(chunk_text)
                
                is_hit = False
                for gold_text in gold_evidence_texts:
                    hit, sim = self.check_retrieved_hit_gold(chunk_text, [gold_text])
                    if hit:
                        is_hit = True
                        break
                
                recall_results[k] = 1.0 if is_hit else 0.0
            else:
                # Coverage Recall: Top-K覆盖了多少比例的Gold Evidence
                gold_hit_flags = [False] * len(gold_evidence_texts)
                
                hits = 0
                for g_idx, gold_text in enumerate(gold_evidence_texts):
                    for chunk in top_k_chunks:
                        chunk_text = chunk.get('text', '') or chunk.get('content', '')
                        chunk_text = self.clean_markdown(chunk_text)
                        is_hit, sim = self.check_retrieved_hit_gold(chunk_text, [gold_text])
                        if is_hit:
                            gold_hit_flags[g_idx] = True
                            hits += 1
                            break
                
                recall_results[k] = hits / len(gold_evidence_texts)
        
        return recall_results
    
    def calculate_mrr(self, reranked_chunks: List[Dict], 
                      gold_evidence_texts: List[str]) -> float:
        """
        基于内容匹配计算MRR
        找到第一个命中任意Gold Evidence的chunk的排名
        """
        if not gold_evidence_texts:
            return None
        
        for rank, chunk in enumerate(reranked_chunks, start=1):
            chunk_text = chunk.get('text', '') or chunk.get('content', '')
            # 清理markdown格式，提高与Gold Evidence的匹配度
            chunk_text = self.clean_markdown(chunk_text)
            is_hit, sim = self.check_retrieved_hit_gold(chunk_text, gold_evidence_texts)
            if is_hit:
                return 1.0 / rank
        
        return 0.0
    
    def extract_citations_from_answer(self, answer: str) -> List[Dict]:
        """
        从 LLM 生成的回答中提取引用信息
        支持格式：
        - 来自: [论文名, 页码, chunk_id]
        - 来自: [论文名 (第X页, chunk_id)]
        """
        citations = []
        
        # 匹配 "来自: [xxx]" 格式
        pattern = r'来自:\s*\[([^\]]+)\]'
        matches = re.findall(pattern, answer)
        
        for match in matches:
            # 尝试匹配括号格式 "论文名 (第X页, chunk_id)"
            bracket_pattern = r'([^\(]+)\s*\(([^,]+),\s*([^\)]+)\)'
            bracket_match = re.search(bracket_pattern, match)
            if bracket_match:
                title = bracket_match.group(1).strip()
                page = bracket_match.group(2).strip()
                chunk_id = bracket_match.group(3).strip()
                # 组合成完整的数据库ID: 论文名_chunk_id
                db_id = f"{title}_{chunk_id}" if not chunk_id.startswith(title.split('_')[0]) else chunk_id
                citations.append({
                    'title': title,
                    'page': page,
                    'chunk_id': chunk_id,
                    'db_id': db_id  # 用于数据库查询的完整ID
                })
                continue
            
            # 逗号分隔格式 "论文名, 页码, chunk_id"
            parts = [p.strip() for p in match.split(',')]
            if len(parts) >= 3:
                title = parts[0]
                page = parts[1]
                chunk_id = parts[2]
                # 组合成完整的数据库ID: 论文名_chunk_id
                db_id = f"{title}_{chunk_id}" if not chunk_id.startswith(title.split('_')[0]) else chunk_id
                citations.append({
                    'title': title,
                    'page': page,
                    'chunk_id': chunk_id,
                    'db_id': db_id  # 用于数据库查询的完整ID
                })
        
        return citations

    def clean_markdown(self, text: str) -> str:
        """
        清理markdown格式，保留纯文本内容
        """
        import re
        
        # 移除代码块
        text = re.sub(r'```[\s\S]*?```', '', text)
        # 移除行内代码
        text = re.sub(r'`([^`]+)`', r'\1', text)
        # 移除图片标记
        text = re.sub(r'!\[([^\]]*)\]\([^)]+\)', r'\1', text)
        # 移除链接，保留文本
        text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)
        # 移除标题标记
        text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
        # 移除粗体/斜体标记
        text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
        text = re.sub(r'\*([^*]+)\*', r'\1', text)
        text = re.sub(r'__([^_]+)__', r'\1', text)
        text = re.sub(r'_([^_]+)_', r'\1', text)
        # 移除列表标记
        text = re.sub(r'^\s*[-*+]\s+', '', text, flags=re.MULTILINE)
        text = re.sub(r'^\s*\d+\.\s+', '', text, flags=re.MULTILINE)
        # 移除引用标记
        text = re.sub(r'^\s*>\s+', '', text, flags=re.MULTILINE)
        # 移除水平线
        text = re.sub(r'^\s*[-*_]{3,}\s*$', '', text, flags=re.MULTILINE)
        # 移除多余的空行（保留单个换行）
        text = re.sub(r'\n{3,}', '\n\n', text)
        # 去除首尾空白
        text = text.strip()
        
        return text
        
    def get_chunk_text_from_db(self, chunk_id: str) -> str:
        """
        从向量数据库中根据chunk_id查询原文
        
        :param chunk_id: chunk的唯一标识
        :return: chunk的原始文本（markdown格式）
        """
        if self.chroma_collection is None:
            return ''
        
        try:
            # 使用 ChromaDB 的 get 方法根据 ID 查询
            result = self.chroma_collection.get(ids=[chunk_id])
            if result and result.get('documents') and len(result['documents']) > 0:
                raw_text = result['documents'][0]
                # 清理markdown格式
                return self.clean_markdown(raw_text)
        except Exception as e:
            print(f"⚠️ 从向量数据库查询 chunk_id={chunk_id} 失败: {e}")
        
        return ''
    
    def calculate_citation_accuracy(self, predicted_answer: str,
                                     predicted_sources: List[Dict],
                                     gold_evidence_texts: List[str],
                                     retrieved_chunks: List[Dict]) -> Tuple[float, bool]:
        """
        计算引用准确率（F1 Score）
        从向量数据库中根据chunk_id查询原文，然后与Gold Evidence进行语义匹配
        """
        extracted_citations = self.extract_citations_from_answer(predicted_answer)
        
        if not extracted_citations:
            return 0.0, False
            
        if not gold_evidence_texts:
            return 0.0, True 
            
        tp, fp = 0, 0
        matched_gold_indices = set() 
        
        for citation in extracted_citations:
            chunk_id = citation.get('chunk_id', '')
            db_id = citation.get('db_id', chunk_id)  # 使用完整的数据库ID
            
            # 策略1: 优先从向量数据库中查询原文（最准确）
            src_text = self.get_chunk_text_from_db(db_id)
            
            # 策略2: 如果向量数据库查询失败，尝试用原始chunk_id查询
            if not src_text and db_id != chunk_id:
                src_text = self.get_chunk_text_from_db(chunk_id)
            
            # 策略3: 如果向量数据库查询失败，尝试从检索结果中找
            if not src_text:
                for chunk in retrieved_chunks:
                    if chunk.get('chunk_id', '') == chunk_id or chunk.get('chunk_id', '') == db_id:
                        src_text = chunk.get('text', '')
                        break
            
            # 策略3: 如果还没找到，尝试从 predicted_sources 中获取
            if not src_text and predicted_sources:
                for source in predicted_sources:
                    if source.get('chunk_id', '') == chunk_id:
                        src_text = source.get('text', '')
                        break
            
            # 如果实在拿不到引用的源文本，直接判为错误引用 (FP)
            if not src_text:
                fp += 1
                continue
            # ----------------------------------------

            hit_idx, sim = self.check_citation_hit_gold(src_text, gold_evidence_texts)
            if hit_idx != -1:
                if hit_idx not in matched_gold_indices:
                    tp += 1
                    matched_gold_indices.add(hit_idx)
                else:
                    fp += 1  # 重复引用同一个 Gold
            else:
                fp += 1      # 引用了无关内容

        fn = len(gold_evidence_texts) - len(matched_gold_indices) 

        precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        
        f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0

        return round(f1_score, 4), True    
    
    def evaluate_single_question(self, question: str, question_type: str,
                                  gold_answer: str, gold_evidence_str: str,
                                  predicted_answer: str, predicted_sources: List[Dict],
                                  reranked_chunks: List[Dict],
                                  manual_score: float = None,
                                  ragas_metrics: Dict = None) -> Dict:
        """
        评估单个问题（基于内容匹配）
        
        :param reranked_chunks: 重排后的chunk列表（用于生成回答的上下文）
        :param manual_score: 人工评分（0-1），如果提供则使用人工评分
        :param ragas_metrics: RAGas 评测结果（包含 tp/fp/fn），如果提供则直接使用
        """
        # 解析Gold Evidence
        gold_evidence_texts = self.parse_gold_evidence(gold_evidence_str)
        
        # 判断是否为不可回答问题
        is_unanswerable = (question_type == "不可回答" or 
                          not gold_evidence_texts or 
                          (len(gold_evidence_texts) == 1 and 
                           any(keyword in gold_evidence_texts[0].lower() 
                              for keyword in ['无法回答', '不可回答', '无', 'none', 'unanswerable'])))
        
        if is_unanswerable:
            # 不可回答类问题：只计算回答正确率
            # 判断模型是否正确拒答
            is_correct_refusal = ("无法回答此问题" in predicted_answer or 
                                 "不可回答" in predicted_answer or
                                 "无法提供" in predicted_answer or
                                 "没有足够信息" in predicted_answer)
            
            # 如果Gold Answer也是拒答，且模型也拒答 → TP
            gold_is_refusal = ("无法回答" in str(gold_answer) or 
                              "不可回答" in str(gold_answer) or
                              "无法提供" in str(gold_answer))
            
            if gold_is_refusal and is_correct_refusal:
                # 正确拒答：TP=1, FP=0, FN=0
                correctness_score = 1.0
                tp, fp, fn = 1, 0, 0
            elif gold_is_refusal and not is_correct_refusal:
                # 模型强行回答（幻觉）：TP=0, FP=1, FN=0
                correctness_score = 0.0
                tp, fp, fn = 0, 1, 0
            elif not gold_is_refusal and is_correct_refusal:
                # 模型错误拒答：TP=0, FP=0, FN=1
                correctness_score = 0.0
                tp, fp, fn = 0, 0, 1
            else:
                # 两者都有答案，用LLM评分
                correctness_result = self.calculate_ragas_style_correctness(
                    question, str(gold_answer), predicted_answer
                )
                correctness_score = correctness_result['f1_score']
                tp, fp, fn = correctness_result['tp'], correctness_result['fp'], correctness_result['fn']
            
            return {
                "question": question,
                "question_type": question_type,
                "gold_answer": gold_answer,
                "predicted_answer": predicted_answer,
                "recall_at_1": None,  # 不可回答问题不计算Recall
                "recall_at_3": None,
                "recall_at_5": None,
                "mrr": None,  # 不可回答问题不计算MRR
                "citation_accuracy": None,  # 不可回答问题不计算引用准确率
                "answer_correctness": manual_score if manual_score is not None else correctness_score,
                "is_unanswerable": True,
                "is_correct_refusal": is_correct_refusal,
                "tp": tp,
                "fp": fp,
                "fn": fn
            }
        
        # 可回答问题：正常计算所有指标
        # 计算重排后的Recall@k
        recall_at_k = self.calculate_recall_at_k(
            reranked_chunks, gold_evidence_texts, k_values=[1, 3, 5]
        )
        
        # 计算重排后的MRR
        mrr = self.calculate_mrr(reranked_chunks, gold_evidence_texts)
        
        # 计算引用准确率（基于重排后的结果）
        citation_accuracy, has_citation = self.calculate_citation_accuracy(
            predicted_answer, predicted_sources, gold_evidence_texts, reranked_chunks
        )
        
        # 计算回答正确率（TP/FP/FN）
        if ragas_metrics is not None:
            # 优先使用传入的 RAGas 评测结果
            tp = ragas_metrics['tp']
            fp = ragas_metrics['fp']
            fn = ragas_metrics['fn']
        elif manual_score is not None:
            # 如果只有评分没有详细指标，使用默认值
            tp, fp, fn = 0, 0, 0
        else:
            # 都没有，重新计算
            correctness_result = self.calculate_ragas_style_correctness(
                question, str(gold_answer), predicted_answer
            )
            tp, fp, fn = correctness_result['tp'], correctness_result['fp'], correctness_result['fn']
        
        return {
            "question": question,
            "question_type": question_type,
            "gold_answer": gold_answer,
            "predicted_answer": predicted_answer,
            "recall_at_1": recall_at_k[1],
            "recall_at_3": recall_at_k[3],
            "recall_at_5": recall_at_k[5],
            "mrr": mrr,
            "citation_accuracy": citation_accuracy,
            "answer_correctness": manual_score,
            "is_unanswerable": False,
            "tp": tp,
            "fp": fp,
            "fn": fn
        }
    
    def run_evaluation(self, rag_pipeline_func, output_file: str = "evaluation_results.xlsx",
                       manual_scoring: bool = True):
        """
        运行完整评测
        """
        total = len(self.test_df)
        print(f"\n🚀 开始评测，共 {total} 个问题...")
        print(f"   使用内容匹配模式（相似度阈值={self.similarity_threshold}）\n")
        all_results = []
        
        for idx, row in enumerate(self.test_df.iterrows()):
            idx, row = idx, row[1]
            question = row['问题']
            question_type = row['question_type']
            gold_answer = row['参考答案']
            gold_evidence = row['Gold Evidence']
            
            # 打印进度条
            progress = (idx + 1) / total
            bar_length = 40
            filled = int(bar_length * progress)
            bar = '█' * filled + '░' * (bar_length - filled)
            percent = progress * 100
            
            print(f"\n{'='*80}")
            print(f"📊 进度: [{bar}] {percent:.1f}% ({idx+1}/{total})")
            print(f"{'='*80}")
            print(f"❓ 问题 {idx+1}/{total}: {question[:80]}{'...' if len(question) > 80 else ''}")
            print(f"{'='*80}")
            
            try:
                result = rag_pipeline_func(question)
                
                # 自动评分或人工评分
                # 不可回答问题跳过 LLM 评分，由 evaluate_single_question 内部处理
                is_unanswerable = (question_type == "不可回答" or 
                                  pd.isna(gold_evidence) or str(gold_evidence).strip() == "")
                
                if self.auto_scoring and not is_unanswerable:
                    print(f"\n{'='*80}")
                    print(f"问题 {idx+1}/{len(self.test_df)}: {question}")
                    print(f"{'='*80}")
                    print(f"\n📝 模型回答:")
                    print(result['answer'])
                    print(f"\n📖 参考答案:")
                    print(gold_answer)
                    print(f"\n🤖 LLM 自动评分中...")
                    '''
                    manual_score = self.score_answer_with_llm(
                        question, result['answer'], gold_answer
                    )
                    '''
                    # 替换为 RAGas 风格评测：
                    print(f"\n🧠 正在进行 RAGas 风格知识点拆解评测...")
                    ragas_metrics = self.calculate_ragas_style_correctness(
                        question, gold_answer, result['answer']
                    )
                    manual_score = ragas_metrics["f1_score"]

                    print(f"✅ 知识点统计: 命中(TP)={ragas_metrics['tp']}, 幻觉(FP)={ragas_metrics['fp']}, 遗漏(FN)={ragas_metrics['fn']}")
                    print(f"✅ RAGas F1 正确率: {manual_score}")

                    if manual_score is not None:
                        print(f"✅ 自动评分: {manual_score:.2f}")
                    else:
                        print("⚠️  自动评分失败，跳过评分")
                elif is_unanswerable:
                    # 不可回答问题，跳过评分
                    manual_score = None
                    ragas_metrics = None
                elif manual_scoring:
                    print(f"\n{'='*80}")
                    print(f"问题 {idx+1}/{len(self.test_df)}: {question}")
                    print(f"{'='*80}")
                    print(f"\n📝 模型回答:")
                    print(result['answer'])
                    print(f"\n📖 参考答案:")
                    print(gold_answer)
                    print(f"\n{'='*80}")
                    
                    while True:
                        try:
                            score_input = input("请打分 (0-1, 或 s 跳过): ").strip()
                            if score_input.lower() == 's':
                                manual_score = None
                                break
                            manual_score = float(score_input)
                            if 0 <= manual_score <= 1:
                                break
                            else:
                                print("⚠️  请输入 0-1 之间的数字")
                        except ValueError:
                            print("⚠️  请输入有效的数字")
                else:
                    manual_score = None
                    ragas_metrics = None
                
                metrics = self.evaluate_single_question(
                    question=question,
                    question_type=question_type,
                    gold_answer=gold_answer,
                    gold_evidence_str=gold_evidence,
                    predicted_answer=result['answer'],
                    predicted_sources=result.get('sources', []),
                    reranked_chunks=result.get('context_docs', []),
                    manual_score=manual_score,
                    ragas_metrics=ragas_metrics if (self.auto_scoring and not is_unanswerable) else None
                )
                
                all_results.append(metrics)
                
                # 立即追加写入文件（方便及时查看）
                self._append_to_excel(metrics, output_file)
                
                # 立即打印当前问题的评测指标
                print(f"\n{'='*80}")
                print(f"📊 问题 {idx+1} 评测指标（重排后）:")
                print(f"{'='*80}")
                print(f"  Recall@1:  {metrics['recall_at_1']:.4f}" if metrics['recall_at_1'] is not None else "  Recall@1:  N/A")
                print(f"  Recall@3:  {metrics['recall_at_3']:.4f}" if metrics['recall_at_3'] is not None else "  Recall@3:  N/A")
                print(f"  Recall@5:  {metrics['recall_at_5']:.4f}" if metrics['recall_at_5'] is not None else "  Recall@5:  N/A")
                print(f"  MRR:       {metrics['mrr']:.4f}" if metrics['mrr'] is not None else "  MRR:       N/A")
                if metrics.get('citation_accuracy') is not None:
                    print(f"  引用准确率: {metrics['citation_accuracy']:.4f}")
                else:
                    print(f"  引用准确率: N/A")
                if metrics['answer_correctness'] is not None:
                    print(f"  回答正确率:   {metrics['answer_correctness']:.2f}")
                print(f"{'='*80}")
                
            except Exception as e:
                print(f"\n❌ 问题 '{question[:50]}...' 评测失败: {e}")
                all_results.append({
                    "question": question,
                    "question_type": question_type,
                    "gold_answer": "",
                    "predicted_answer": "",
                    "error": str(e),
                    "recall_at_1": None, "recall_at_3": None,
                    "recall_at_5": None,
                    "mrr": None, "citation_accuracy": None,
                    "answer_correctness": None,
                    "is_unanswerable": False,
                    "tp": None, "fp": None, "fn": None
                })
        
        # 保存结果（已经逐题追加，这里再保存一份完整的用于备份）
        results_df = pd.DataFrame(all_results)
        results_df.to_excel(output_file.replace('.xlsx', '_final.xlsx'), index=False)
        
        # 在输出文件末尾添加平均值行
        self._append_averages_to_excel(results_df, output_file)
        
        # 打印汇总
        self._print_summary(results_df)
        
        return results_df
    
    def _append_averages_to_excel(self, results_df: pd.DataFrame, output_file: str):
        """在Excel文件末尾添加平均值行（使用Excel公式，方便后续调整）"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("⚠️ 未安装 openpyxl，将使用普通平均值写入")
            self._append_averages_simple(results_df, output_file)
            return
        
        # 先保存普通Excel文件（包含数据）
        results_df.to_excel(output_file, index=False)
        
        # 使用 openpyxl 添加公式行
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 获取列名对应的字母
        col_names = list(results_df.columns)
        col_map = {name: idx + 1 for idx, name in enumerate(col_names)}  # 列号从1开始
        
        total_rows = len(results_df) + 1  # +1 因为Excel行号从1开始，且有表头
        
        # 需要计算平均值的指标列
        metrics_cols = ['recall_at_1', 'recall_at_3', 'recall_at_5', 'mrr', 
                       'citation_accuracy', 'answer_correctness', 'tp', 'fp', 'fn']
        
        # 获取列号
        def get_col_letter(col_name):
            """将列号转换为Excel列字母"""
            if col_name not in col_map:
                return None
            col_num = col_map[col_name]
            letter = ''
            while col_num > 0:
                col_num, remainder = divmod(col_num - 1, 26)
                letter = chr(65 + remainder) + letter
            return letter
        
        # 添加总体平均值行
        avg_row_num = total_rows + 2  # 空一行后添加
        ws.cell(row=avg_row_num, column=col_map.get('question', 1), value='【总体平均值】')
        
        for metric in metrics_cols:
            col_letter = get_col_letter(metric)
            if col_letter:
                # 使用 AVERAGEIF 忽略空值
                formula = f'=AVERAGEIF({col_letter}2:{col_letter}{total_rows},">0")'
                ws.cell(row=avg_row_num, column=col_map[metric], value=formula)
        
        # 按问题类型添加平均值行
        current_row = avg_row_num + 1
        for q_type in self.question_types:
            type_df = results_df[results_df['question_type'] == q_type]
            if len(type_df) == 0:
                continue
            
            ws.cell(row=current_row, column=col_map.get('question', 1), value=f'【{q_type}平均值】')
            ws.cell(row=current_row, column=col_map.get('question_type', 2), value=q_type)
            
            for metric in metrics_cols:
                col_letter = get_col_letter(metric)
                if col_letter:
                    formula = f'=AVERAGEIF({col_letter}2:{col_letter}{total_rows},">0")'
                    ws.cell(row=current_row, column=col_map[metric], value=formula)
            
            current_row += 1
        
        wb.save(output_file)
        print(f"\n✅ 平均值公式行已追加到: {output_file}")
    
    def _append_averages_simple(self, results_df: pd.DataFrame, output_file: str):
        """备用方法：使用普通平均值（无openpyxl时）"""
        # 分离可回答和不可回答问题
        answerable_df = results_df[results_df.get('is_unanswerable', False) == False]
        
        all_avg_rows = []
        
        # 1. 总体平均值
        overall_avg = {
            "question": "【总体平均值】",
            "question_type": "",
            "gold_answer": "",
            "predicted_answer": "",
            "recall_at_1": answerable_df['recall_at_1'].mean() if len(answerable_df) > 0 else None,
            "recall_at_3": answerable_df['recall_at_3'].mean() if len(answerable_df) > 0 else None,
            "recall_at_5": answerable_df['recall_at_5'].mean() if len(answerable_df) > 0 else None,
            "mrr": answerable_df['mrr'].mean() if len(answerable_df) > 0 else None,
            "citation_accuracy": answerable_df['citation_accuracy'].mean() if len(answerable_df) > 0 else None,
            "answer_correctness": results_df['answer_correctness'].mean() if len(results_df) > 0 else None,
            "is_unanswerable": False,
            "tp": results_df['tp'].mean() if 'tp' in results_df.columns and len(results_df) > 0 else None,
            "fp": results_df['fp'].mean() if 'fp' in results_df.columns and len(results_df) > 0 else None,
            "fn": results_df['fn'].mean() if 'fn' in results_df.columns and len(results_df) > 0 else None,
        }
        all_avg_rows.append(overall_avg)
        
        # 2. 按问题类型分类的平均值
        for q_type in self.question_types:
            type_df = results_df[results_df['question_type'] == q_type]
            if len(type_df) == 0:
                continue
            
            type_answerable = type_df[type_df.get('is_unanswerable', False) == False]
            
            type_avg = {
                "question": f"【{q_type}平均值】",
                "question_type": q_type,
                "gold_answer": "",
                "predicted_answer": "",
                "recall_at_1": type_answerable['recall_at_1'].mean() if len(type_answerable) > 0 else None,
                "recall_at_3": type_answerable['recall_at_3'].mean() if len(type_answerable) > 0 else None,
                "recall_at_5": type_answerable['recall_at_5'].mean() if len(type_answerable) > 0 else None,
                "mrr": type_answerable['mrr'].mean() if len(type_answerable) > 0 else None,
                "citation_accuracy": type_answerable['citation_accuracy'].mean() if len(type_answerable) > 0 else None,
                "answer_correctness": type_df['answer_correctness'].mean() if len(type_df) > 0 else None,
                "is_unanswerable": False,
                "tp": type_df['tp'].mean() if 'tp' in type_df.columns and len(type_df) > 0 else None,
                "fp": type_df['fp'].mean() if 'fp' in type_df.columns and len(type_df) > 0 else None,
                "fn": type_df['fn'].mean() if 'fn' in type_df.columns and len(type_df) > 0 else None,
            }
            all_avg_rows.append(type_avg)
        
        # 追加到Excel文件
        avg_df = pd.DataFrame(all_avg_rows)
        if os.path.exists(output_file):
            existing_df = pd.read_excel(output_file)
            combined_df = pd.concat([existing_df, avg_df], ignore_index=True)
        else:
            combined_df = avg_df
        
        combined_df.to_excel(output_file, index=False)
        print(f"\n✅ 平均值行已追加到: {output_file}")
    
    def _append_to_excel(self, metrics: Dict, output_file: str):
        """追加单个问题的评测结果到 Excel 文件"""
        # 将 metrics 转换为 DataFrame
        new_row = pd.DataFrame([metrics])
        
        # 如果文件存在，追加写入；否则创建新文件
        if os.path.exists(output_file):
            existing_df = pd.read_excel(output_file)
            combined_df = pd.concat([existing_df, new_row], ignore_index=True)
        else:
            combined_df = new_row
        
        combined_df.to_excel(output_file, index=False)
    
    def _print_summary(self, results_df: pd.DataFrame):
        """打印评测汇总"""
        print("\n" + "="*80)
        print("📊 评测结果汇总（基于内容匹配）")
        print("="*80)
        
        # 分离可回答和不可回答问题
        answerable_df = results_df[results_df.get('is_unanswerable', False) == False]
        unanswerable_df = results_df[results_df.get('is_unanswerable', False) == True]
        
        print(f"\n【总体指标】（仅可回答问题: {len(answerable_df)}/{len(results_df)}题）")
        for metric in ['recall_at_1', 'recall_at_3', 'recall_at_5', 'mrr', 'citation_accuracy']:
            val = answerable_df[metric].mean()
            if pd.notna(val):
                print(f"  {metric}: {val:.4f}")
        
        # 不可回答类问题的正确拒答率
        if len(unanswerable_df) > 0:
            correct_refusal_rate = unanswerable_df['is_correct_refusal'].mean()
            print(f"\n【不可回答问题】（{len(unanswerable_df)}题）")
            print(f"  正确拒答率: {correct_refusal_rate:.4f}")
        
        print("\n【按问题类型细分】")
        for q_type in self.question_types:
            type_df = results_df[results_df['question_type'] == q_type]
            if len(type_df) == 0:
                continue
            
            # 检查该类型是否包含不可回答问题
            type_unanswerable = type_df[type_df.get('is_unanswerable', False) == True]
            type_answerable = type_df[type_df.get('is_unanswerable', False) == False]
            
            print(f"\n  {q_type} ({len(type_df)}题):")
            
            if len(type_unanswerable) > 0:
                correct_refusal_rate = type_unanswerable['is_correct_refusal'].mean()
                print(f"    不可回答问题: {len(type_unanswerable)}题, 正确拒答率: {correct_refusal_rate:.4f}")
            
            if len(type_answerable) > 0:
                # 输出检索指标
                for metric in ['recall_at_1', 'recall_at_3', 'recall_at_5', 'mrr', 'citation_accuracy']:
                    val = type_answerable[metric].mean()
                    if pd.notna(val):
                        print(f"    {metric}: {val:.4f}")
                
                # 输出回答正确率
                correctness_val = type_answerable['answer_correctness'].mean()
                if pd.notna(correctness_val):
                    print(f"    answer_correctness: {correctness_val:.4f}")
            elif len(type_unanswerable) == 0:
                print("    (无有效评测数据)")
        
        print("\n" + "="*80)


# ===== 使用示例 =====
if __name__ == "__main__":
    from rag_pipeline import ask, RAGModels
    import chromadb
    from chromadb.config import Settings
    
    models = RAGModels()
    chroma_client = chromadb.PersistentClient(
        path="E:\\rag_project\\code\\vector_db\\bge-m3\\chunk_512",
        settings=Settings(anonymized_telemetry=False),
    )
    collection = chroma_client.get_collection("rag_papers_512")
    
    def rag_pipeline_wrapper(question: str) -> Dict:
        result = ask(question, models, collection)
        return {
            'answer': result.get('answer', ''),
            'sources': result.get('sources', []),
            'context_docs': result.get('context_docs', [])  # 重排后的结果（用于评估）
        }
    
    evaluator = ContentBasedRAGEvaluator(
        test_excel="测试集.xlsx",
        similarity_threshold=0.70,  # 可根据实际情况调整
        auto_scoring=True,  # 启用 LLM 自动评分
        chroma_collection=collection  # 传入向量数据库，用于引用验证
    )
    
    # 复用 RAG pipeline 中的 LLM 作为评分模型
    evaluator.scoring_model = models
    
    results_df = evaluator.run_evaluation(
        rag_pipeline_func=rag_pipeline_wrapper,
        output_file="evaluation_results.xlsx"
    )