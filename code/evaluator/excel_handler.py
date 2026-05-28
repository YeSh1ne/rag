import pandas as pd
import os
from typing import Dict, List


def append_to_excel(metrics: Dict, output_file: str):
    """追加单个问题的评测结果到 Excel 文件"""
    # 将 metrics 转换为 DataFrame
    new_row = pd.DataFrame([metrics])
    
    # 如果文件存在，追加写入；否则创建新文件
    if os.path.exists(output_file):
        existing_df = pd.read_excel(output_file)
        combined_df = pd.concat([existing_df, new_row], ignore_index=True)
    else:
        combined_df = new_row
    
    combined_df.to_excel(output_file, index=False)


def append_averages_to_excel(results_df: pd.DataFrame, output_file: str, question_types: List[str]):
    """在Excel文件末尾添加平均值行（使用Excel公式，方便后续调整）"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️ 未安装 openpyxl，将使用普通平均值写入")
        _append_averages_simple(results_df, output_file, question_types)
        return
    
    # 先保存普通Excel文件（包含数据）
    results_df.to_excel(output_file, index=False)
    
    # 使用 openpyxl 添加公式行
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 获取列名对应的字母
    col_names = list(results_df.columns)
    col_map = {name: idx + 1 for idx, name in enumerate(col_names)}  # 列号从1开始
    
    total_rows = len(results_df) + 1  # +1 因为Excel行号从1开始，且有表头
    
    # 需要计算平均值的指标列
    metrics_cols = ['recall_at_1', 'recall_at_3', 'recall_at_5', 'mrr', 
                   'citation_accuracy', 'answer_correctness', 'tp', 'fp', 'fn']
    
    # 获取列号
    def get_col_letter(col_name):
        """将列号转换为Excel列字母"""
        if col_name not in col_map:
            return None
        col_num = col_map[col_name]
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    # 添加总体平均值行
    avg_row_num = total_rows + 2  # 空一行后添加
    ws.cell(row=avg_row_num, column=col_map.get('question', 1), value='【总体平均值】')
    
    for metric in metrics_cols:
        col_letter = get_col_letter(metric)
        if col_letter:
            # 使用 AVERAGEIF 忽略空值
            formula = f'=AVERAGEIF({col_letter}2:{col_letter}{total_rows},">0")'
            ws.cell(row=avg_row_num, column=col_map[metric], value=formula)
    
    # 按问题类型添加平均值行
    current_row = avg_row_num + 1
    for q_type in question_types:
        type_df = results_df[results_df['question_type'] == q_type]
        if len(type_df) == 0:
            continue
        
        ws.cell(row=current_row, column=col_map.get('question', 1), value=f'【{q_type}平均值】')
        ws.cell(row=current_row, column=col_map.get('question_type', 2), value=q_type)
        
        for metric in metrics_cols:
            col_letter = get_col_letter(metric)
            if col_letter:
                formula = f'=AVERAGEIF({col_letter}2:{col_letter}{total_rows},">0")'
                ws.cell(row=current_row, column=col_map[metric], value=formula)
        
        current_row += 1
    
    wb.save(output_file)
    print(f"\n✅ 平均值公式行已追加到: {output_file}")


def _append_averages_simple(results_df: pd.DataFrame, output_file: str, question_types: List[str]):
    """备用方法：使用普通平均值（无openpyxl时）"""
    # 分离可回答和不可回答问题
    answerable_df = results_df[results_df.get('is_unanswerable', False) == False]
    
    all_avg_rows = []
    
    # 1. 总体平均值
    overall_avg = {
        "question": "【总体平均值】",
        "question_type": "",
        "gold_answer": "",
        "predicted_answer": "",
        "recall_at_1": answerable_df['recall_at_1'].mean() if len(answerable_df) > 0 else None,
        "recall_at_3": answerable_df['recall_at_3'].mean() if len(answerable_df) > 0 else None,
        "recall_at_5": answerable_df['recall_at_5'].mean() if len(answerable_df) > 0 else None,
        "mrr": answerable_df['mrr'].mean() if len(answerable_df) > 0 else None,
        "citation_accuracy": answerable_df['citation_accuracy'].mean() if len(answerable_df) > 0 else None,
        "answer_correctness": results_df['answer_correctness'].mean() if len(results_df) > 0 else None,
        "is_unanswerable": False,
        "tp": results_df['tp'].mean() if 'tp' in results_df.columns and len(results_df) > 0 else None,
        "fp": results_df['fp'].mean() if 'fp' in results_df.columns and len(results_df) > 0 else None,
        "fn": results_df['fn'].mean() if 'fn' in results_df.columns and len(results_df) > 0 else None,
    }
    all_avg_rows.append(overall_avg)
    
    # 2. 按问题类型分类的平均值
    for q_type in question_types:
        type_df = results_df[results_df['question_type'] == q_type]
        if len(type_df) == 0:
            continue
        
        type_answerable = type_df[type_df.get('is_unanswerable', False) == False]
        
        type_avg = {
            "question": f"【{q_type}平均值】",
            "question_type": q_type,
            "gold_answer": "",
            "predicted_answer": "",
            "recall_at_1": type_answerable['recall_at_1'].mean() if len(type_answerable) > 0 else None,
            "recall_at_3": type_answerable['recall_at_3'].mean() if len(type_answerable) > 0 else None,
            "recall_at_5": type_answerable['recall_at_5'].mean() if len(type_answerable) > 0 else None,
            "mrr": type_answerable['mrr'].mean() if len(type_answerable) > 0 else None,
            "citation_accuracy": type_answerable['citation_accuracy'].mean() if len(type_answerable) > 0 else None,
            "answer_correctness": type_df['answer_correctness'].mean() if len(type_df) > 0 else None,
            "is_unanswerable": False,
            "tp": type_df['tp'].mean() if 'tp' in type_df.columns and len(type_df) > 0 else None,
            "fp": type_df['fp'].mean() if 'fp' in type_df.columns and len(type_df) > 0 else None,
            "fn": type_df['fn'].mean() if 'fn' in type_df.columns and len(type_df) > 0 else None,
        }
        all_avg_rows.append(type_avg)
    
    # 追加到Excel文件
    avg_df = pd.DataFrame(all_avg_rows)
    if os.path.exists(output_file):
        existing_df = pd.read_excel(output_file)
        combined_df = pd.concat([existing_df, avg_df], ignore_index=True)
    else:
        combined_df = avg_df
    
    combined_df.to_excel(output_file, index=False)
    print(f"\n✅ 平均值行已追加到: {output_file}")


def print_summary(results_df: pd.DataFrame, question_types: List[str]):
    """打印评测汇总"""
    print("\n" + "="*80)
    print("📊 评测结果汇总（基于内容匹配）")
    print("="*80)
    
    # 分离可回答和不可回答问题
    answerable_df = results_df[results_df.get('is_unanswerable', False) == False]
    unanswerable_df = results_df[results_df.get('is_unanswerable', False) == True]
    
    print(f"\n【总体指标】（仅可回答问题: {len(answerable_df)}/{len(results_df)}题）")
    for metric in ['recall_at_1', 'recall_at_3', 'recall_at_5', 'mrr', 'citation_accuracy']:
        val = answerable_df[metric].mean()
        if pd.notna(val):
            print(f"  {metric}: {val:.4f}")
    
    # 不可回答类问题的正确拒答率
    if len(unanswerable_df) > 0:
        correct_refusal_rate = unanswerable_df['is_correct_refusal'].mean()
        print(f"\n【不可回答问题】（{len(unanswerable_df)}题）")
        print(f"  正确拒答率: {correct_refusal_rate:.4f}")
    
    print("\n【按问题类型细分】")
    for q_type in question_types:
        type_df = results_df[results_df['question_type'] == q_type]
        if len(type_df) == 0:
            continue
        
        # 检查该类型是否包含不可回答问题
        type_unanswerable = type_df[type_df.get('is_unanswerable', False) == True]
        type_answerable = type_df[type_df.get('is_unanswerable', False) == False]
        
        print(f"\n  {q_type} ({len(type_df)}题):")
        
        if len(type_unanswerable) > 0:
            correct_refusal_rate = type_unanswerable['is_correct_refusal'].mean()
            print(f"    不可回答问题: {len(type_unanswerable)}题, 正确拒答率: {correct_refusal_rate:.4f}")
        
        if len(type_answerable) > 0:
            # 输出检索指标
            for metric in ['recall_at_1', 'recall_at_3', 'recall_at_5', 'mrr', 'citation_accuracy']:
                val = type_answerable[metric].mean()
                if pd.notna(val):
                    print(f"    {metric}: {val:.4f}")
            
            # 输出回答正确率
            correctness_val = type_answerable['answer_correctness'].mean()
            if pd.notna(correctness_val):
                print(f"    answer_correctness: {correctness_val:.4f}")
        elif len(type_unanswerable) == 0:
            print("    (无有效评测数据)")
    
    print("\n" + "="*80)